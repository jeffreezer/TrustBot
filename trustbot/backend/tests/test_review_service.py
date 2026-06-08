"""Review-workflow pure logic — DB-free tests.

The DB round-trip (status writes, audit entries, org-scoping) is verified end-to-end
against the live stack; here we pin the pure pieces: the action→status transitions and
the export serialization that carries approval status.
"""
import io

import openpyxl
import pytest

from app.questionnaires.service import (
    ACTION_STATUS,
    EXPORT_COLUMNS,
    REVIEW_ACTIONS,
    next_review_status,
    rows_to_csv,
    rows_to_xlsx,
)

SAMPLE_ROWS = [
    {
        "id": "Q01", "domain": "Encryption", "question": "Encrypted at rest?",
        "outcome": "attested", "review_status": "approved",
        "needs_human_review": "false", "confidence": "high",
        "short_answer": "Yes. AES-256.", "answer": "Yes. AES-256 at rest.",
        "documents": "", "remediation": "", "evidence": "Security Whitepaper",
        "freshness": "current", "reviewer": "alice",
    },
    {
        "id": "Q09", "domain": "Compliance", "question": "FedRAMP authorized?",
        "outcome": "needs_input", "review_status": "pending",
        "needs_human_review": "true", "confidence": "none",
        "short_answer": "", "answer": "", "documents": "", "remediation": "",
        "evidence": "", "freshness": "unknown", "reviewer": "",
    },
]


def test_action_status_covers_all_actions():
    assert REVIEW_ACTIONS == {
        "approve", "edit", "reject", "request_evidence", "save_to_library"
    }
    assert ACTION_STATUS["approve"] == "approved"
    assert ACTION_STATUS["edit"] == "edited"
    assert ACTION_STATUS["reject"] == "rejected"
    assert ACTION_STATUS["request_evidence"] == "needs_evidence"
    # Saving to the library implies acceptance.
    assert ACTION_STATUS["save_to_library"] == "approved"


def test_next_review_status_rejects_unknown_action():
    with pytest.raises(ValueError):
        next_review_status("delete_everything")


def test_csv_export_has_status_columns_and_rows():
    body = rows_to_csv(SAMPLE_ROWS).decode("utf-8-sig")
    header = body.splitlines()[0]
    assert "review_status" in header
    assert "needs_human_review" in header
    # Approval status is carried, so nothing reads "final" unless approved.
    assert "approved" in body
    assert "pending" in body
    assert "FedRAMP authorized?" in body


def test_csv_export_header_matches_columns():
    body = rows_to_csv(SAMPLE_ROWS).decode("utf-8-sig")
    assert body.splitlines()[0].split(",") == EXPORT_COLUMNS


def test_xlsx_export_round_trips():
    blob = rows_to_xlsx(SAMPLE_ROWS)
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    assert list(values[0]) == EXPORT_COLUMNS
    assert values[1][EXPORT_COLUMNS.index("review_status")] == "approved"
    assert values[2][EXPORT_COLUMNS.index("review_status")] == "pending"


def test_export_tolerates_missing_keys():
    # A row missing optional keys must not blow up serialization.
    body = rows_to_csv([{"question": "Only a question?"}]).decode("utf-8-sig")
    assert "Only a question?" in body


# --- CSV / formula injection guard (audit F2) -------------------------------

_FORMULA_QUESTION = "=cmd|'/c calc'!A1"


def test_csv_export_neutralizes_formula_injection():
    # Attacker-controlled question text starting with '=' must export as inert text (leading ').
    body = rows_to_csv([{"question": _FORMULA_QUESTION, "answer": "+SUM(A1:A9)"}]).decode("utf-8-sig")
    # The cell value is quoted so a spreadsheet treats it as text, not a live formula.
    assert "'=cmd" in body
    assert "'+SUM" in body
    # A bare (unquoted) leading '=' formula must NOT appear.
    assert "\n=cmd" not in body and ",=cmd" not in body


def test_xlsx_export_neutralizes_formula_injection():
    blob = rows_to_xlsx([{"question": _FORMULA_QUESTION, "short_answer": "@WEBSERVICE(...)"}])
    ws = openpyxl.load_workbook(io.BytesIO(blob)).active
    q_col = EXPORT_COLUMNS.index("question") + 1  # 1-based; row 2 is the first data row
    sa_col = EXPORT_COLUMNS.index("short_answer") + 1
    assert ws.cell(row=2, column=q_col).value == "'" + _FORMULA_QUESTION
    assert ws.cell(row=2, column=sa_col).value == "'@WEBSERVICE(...)"


def test_export_leaves_benign_values_unmodified():
    body = rows_to_csv([{"question": "Do you encrypt at rest?", "confidence": "high"}]).decode("utf-8-sig")
    assert "Do you encrypt at rest?" in body
    assert "'Do you" not in body  # no spurious quoting of safe text
